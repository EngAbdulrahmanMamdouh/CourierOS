import json
from io import BytesIO
from typing import Any

from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import ValidationError

from app.crud import import_job as import_job_crud
from app.crud.shipment import bulk_create_shipments
from app.schemas.shipment import ShipmentCreate


def create_import_job_record(db, file_name: str, current_user, total_rows: int = 0, report: str | None = None):
    return import_job_crud.create_import_job(
        db=db,
        file_name=file_name,
        uploaded_by=current_user.id,
        company_id=getattr(current_user, "company_id", None),
        status="Pending",
        total_rows=total_rows,
        imported_rows=0,
        failed_rows=0,
        report=report,
    )


def _normalize_column_name(column_name: str) -> str:
    return str(column_name or "").strip().lower().replace(" ", "_")


def _create_template_workbook(template_name: str) -> bytes:
    wb = Workbook()
    sheet = wb.active
    if template_name == "shipments":
        headers = ["sender_name", "receiver_name", "receiver_phone", "address", "city", "customer", "cod_amount", "notes"]
        example = ["Alice", "Bob", "01234567890", "1 Main St", "Cairo", "ACME Corp", 50, "Leave at gate"]
    elif template_name == "customers":
        headers = ["full_name", "phone", "email", "company_name", "address", "city", "notes", "is_active"]
        example = ["John Doe", "01234567890", "john@example.com", "ACME Corp", "1 Main St", "Cairo", "Preferred customer", "True"]
    elif template_name == "drivers":
        headers = ["full_name", "phone", "national_id", "license_number", "vehicle_type", "vehicle_plate", "branch_id", "is_active"]
        example = ["Ali Ahmed", "01234567890", "12345678901234", "ABC12345", "Van", "ABC-1234", 1, "True"]
    elif template_name == "branches":
        headers = ["name", "code", "address", "city", "phone", "manager_id", "is_active"]
        example = ["Cairo Branch", "CAI001", "1 Main St", "Cairo", "01234567890", 1, "True"]
    else:
        raise ValueError("Unknown template")

    sheet.append(headers)
    sheet.append(example)

    if template_name in {"customers", "drivers", "branches"}:
        dv = DataValidation(type="list", formula1='"True,False"', allow_blank=True)
        sheet.add_data_validation(dv)
        for row_idx in range(2, 105):
            dv.add(f"{chr(65 + len(headers) - 1)}{row_idx}")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def _get_import_job_report(import_job):
    if not import_job.report:
        return {}
    try:
        return json.loads(import_job.report)
    except Exception:
        return {}


def _create_excel_report(headers: list[str], rows: list[list[Any]]) -> bytes:
    wb = Workbook()
    sheet = wb.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def _generate_error_report_bytes(import_job) -> bytes:
    report = _get_import_job_report(import_job)
    failed_rows = report.get("failed_rows_details", [])
    headers = ["row_number", "error_message"]
    for mapped_key in report.get("failed_columns", []):
        headers.append(mapped_key)

    rows = []
    for failed in failed_rows:
        error_message = "; ".join(failed.get("errors", []))
        row = [failed.get("row_number", ""), error_message]
        mapped_data = failed.get("mapped_data", {})
        for mapped_key in report.get("failed_columns", []):
            row.append(mapped_data.get(mapped_key, ""))
        rows.append(row)

    return _create_excel_report(headers, rows)


def _generate_duplicate_report_bytes(import_job) -> bytes:
    report = _get_import_job_report(import_job)
    duplicate_rows = report.get("duplicate_rows_details", [])
    headers = ["row_number"] + report.get("duplicate_columns", [])
    rows = []
    for dup in duplicate_rows:
        row = [dup.get("row_number", "")]
        mapped_data = dup.get("mapped_data", {})
        for mapped_key in report.get("duplicate_columns", []):
            row.append(mapped_data.get(mapped_key, ""))
        rows.append(row)
    return _create_excel_report(headers, rows)


def _get_import_summary(import_job) -> dict[str, Any]:
    total = import_job.total_rows
    imported = import_job.imported_rows
    failed = import_job.failed_rows
    duplicate = import_job.duplicate_rows
    success_pct = (imported / total * 100) if total else 0.0
    failure_pct = (failed / total * 100) if total else 0.0
    duplicate_pct = (duplicate / total * 100) if total else 0.0
    return {
        "import_job_id": import_job.id,
        "total_rows": total,
        "imported_rows": imported,
        "failed_rows": failed,
        "duplicate_rows": duplicate,
        "success_percentage": round(success_pct, 2),
        "failure_percentage": round(failure_pct, 2),
        "duplicate_percentage": round(duplicate_pct, 2),
        "duration_seconds": import_job.duration_seconds,
        "status": import_job.status,
        "uploaded_by": import_job.uploaded_by,
        "uploaded_time": import_job.uploaded_time,
        "finished_time": import_job.finished_at,
    }


def _detect_known_columns(header_values: list[str]) -> tuple[list[str], list[str]]:
    known_columns = [
        "sender_name",
        "receiver_name",
        "receiver_phone",
        "address",
        "city",
        "customer",
        "cod_amount",
        "notes",
    ]
    normalized = [_normalize_column_name(value) for value in header_values]
    detected = [value for value in normalized if value in known_columns]
    missing_required = [col for col in ["sender_name", "receiver_name", "receiver_phone", "address", "city"] if col not in detected]
    return detected, missing_required


def _map_row(header_values: list[str], row_values: list[Any]) -> dict[str, Any]:
    normalized_headers = [_normalize_column_name(value) for value in header_values]
    mapped = {}
    for name, cell in zip(normalized_headers, row_values):
        if name:
            mapped[name] = cell
    return mapped


def _validate_mapped_data(mapped_data: dict[str, Any]) -> tuple[str, list[str]]:
    errors: list[str] = []
    shipment_payload = {}

    for field in ["sender_name", "receiver_name", "receiver_phone", "address", "city", "customer", "cod_amount", "notes"]:
        if field in mapped_data:
            shipment_payload[field] = mapped_data[field]

    try:
        ShipmentCreate(**shipment_payload)
        status = "valid"
    except ValidationError as exc:
        status = "invalid"
        for err in exc.errors():
            loc = ".".join(str(item) for item in err.get("loc", []))
            msg = err.get("msg", "Invalid value")
            errors.append(f"{loc}: {msg}" if loc else msg)

    return status, errors


def preview_uploaded_workbook(file_bytes: bytes) -> dict[str, Any]:
    try:
        workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ValueError("Invalid Excel file") from exc

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        return {
            "detected_columns": [],
            "missing_required_columns": ["sender_name", "receiver_name", "receiver_phone", "address", "city"],
            "preview_rows": [],
            "total_rows": 0,
        }

    header_values = [str(value).strip() if value is not None else "" for value in rows[0]]
    detected_columns, missing_required_columns = _detect_known_columns(header_values)

    preview_rows = []
    for index, row in enumerate(rows[1:], start=2):
        mapped_data = _map_row(header_values, list(row))
        validation_status, validation_errors = _validate_mapped_data(mapped_data)
        preview_rows.append(
            {
                "row_number": index,
                "validation_status": validation_status,
                "validation_errors": validation_errors,
                "mapped_data": mapped_data,
            }
        )

    return {
        "detected_columns": detected_columns,
        "missing_required_columns": missing_required_columns,
        "preview_rows": preview_rows,
        "total_rows": len(rows) - 1,
    }


def execute_import_workbook(db, file_name: str, file_bytes: bytes, current_user) -> dict[str, Any]:
    from datetime import datetime, timezone

    try:
        workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ValueError("Invalid Excel file") from exc

    started_at = datetime.now(timezone.utc)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    total_rows = max(0, len(rows) - 1)

    header_values = [str(value).strip() if value is not None else "" for value in rows[0]] if rows else []
    detected_columns, missing_required_columns = _detect_known_columns(header_values)

    duplicate_keys: set[tuple[str, str, str, str]] = set()
    validation_errors: list[dict[str, Any]] = []
    valid_shipments: list[ShipmentCreate] = []
    created_shipment_ids: list[int] = []
    duplicate_rows = 0
    failed_rows = 0

    for index, row in enumerate(rows[1:], start=2):
        mapped_data = _map_row(header_values, list(row))
        validation_status, errors = _validate_mapped_data(mapped_data)

        if validation_status != "valid":
            failed_rows += 1
            validation_errors.append({
                "row_number": index,
                "errors": errors,
                "mapped_data": mapped_data,
            })
            continue

        key = (
            str(mapped_data.get("sender_name", "")).strip().lower(),
            str(mapped_data.get("receiver_name", "")).strip().lower(),
            str(mapped_data.get("receiver_phone", "")).strip().lower(),
            str(mapped_data.get("address", "")).strip().lower(),
        )

        if key in duplicate_keys:
            duplicate_rows += 1
            continue

        duplicate_keys.add(key)
        shipment_payload = ShipmentCreate(**{
            field: mapped_data.get(field)
            for field in ["sender_name", "receiver_name", "receiver_phone", "address", "city", "customer", "cod_amount", "notes"]
            if field in mapped_data
        })
        valid_shipments.append(shipment_payload)

    successful_rows = 0
    if valid_shipments:
        shipments = bulk_create_shipments(db, valid_shipments, owner_id=current_user.id, company_id=getattr(current_user, "company_id", 1))
        created_shipment_ids = [shipment.id for shipment in shipments]
        successful_rows = len(shipments)

    finished_at = datetime.now(timezone.utc)
    failed_rows += duplicate_rows

    import_job = import_job_crud.create_import_job(
        db=db,
        file_name=file_name,
        uploaded_by=current_user.id,
        company_id=getattr(current_user, "company_id", None),
        status="Completed",
        total_rows=total_rows,
        imported_rows=successful_rows,
        failed_rows=failed_rows,
        duplicate_rows=duplicate_rows,
        started_at=started_at,
        finished_at=finished_at,
        report=json.dumps({
            "detected_columns": detected_columns,
            "missing_required_columns": missing_required_columns,
            "validation_errors": validation_errors,
            "created_shipment_ids": created_shipment_ids,
        }),
    )

    return {
        "total_rows": total_rows,
        "successful_rows": successful_rows,
        "failed_rows": failed_rows,
        "duplicate_rows": duplicate_rows,
        "validation_errors": validation_errors,
        "created_shipment_ids": created_shipment_ids,
        "execution_time": (finished_at - started_at).total_seconds(),
        "import_job_id": import_job.id,
    }


def store_uploaded_workbook(db, file_name: str, file_bytes: bytes, current_user):
    try:
        workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ValueError("Invalid Excel file") from exc

    sheet = workbook.active
    row_count = sum(
        1
        for row in sheet.iter_rows(values_only=True)
        if any(cell is not None and str(cell).strip() != "" for cell in row)
    )

    workbook_info = {
        "sheet_names": workbook.sheetnames,
        "active_sheet": sheet.title,
        "row_count": row_count,
    }

    return import_job_crud.create_import_job(
        db=db,
        file_name=file_name,
        uploaded_by=current_user.id,
        status="Uploaded",
        total_rows=row_count,
        imported_rows=0,
        failed_rows=0,
        report=json.dumps(workbook_info),
    )


def update_import_job_results(
    db,
    job_id: int,
    current_user,
    status: str,
    imported_rows: int,
    failed_rows: int,
    report: str | None = None,
):
    return import_job_crud.update_import_job(
        db=db,
        job_id=job_id,
        updates={
            "status": status,
            "imported_rows": imported_rows,
            "failed_rows": failed_rows,
            "report": report,
        },
        current_user=current_user,
    )
